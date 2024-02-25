import os
import datetime
import pandas as pd
from .connection import collection
from bson.objectid import ObjectId
from openpyxl.styles import PatternFill

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(BASE_DIR, 'static')

id_filter = {'id':'$_id', 'train_name':1, 'train_number':1, 'arrival_date':1, 'departure_date':1, 'arrival_time':1, 'departure_time':1, 'arrival_timestamp':1, 'departure_timestamp':1, '_id': 0}

def insert_to_db(data):
    document = {
        'train_name': data['train_name'],
        'train_number': data['train_number'],
        'arrival_date': data['arrival_date'],
        'departure_date': data['departure_date'],
        'arrival_time': data['arrival_time'],
        'departure_time': data['departure_time'],
        'arrival_timestamp': data['arrival_timestamp'],
        'departure_timestamp': data['departure_timestamp']
    }
    collection.insert_one(document)


def isValid(data):
    try:
        # Validate date and time formats
        datetime.datetime.strptime(data['arrival_date'], '%Y-%m-%d')
        datetime.datetime.strptime(data['departure_date'], '%Y-%m-%d')
        datetime.datetime.strptime(data['arrival_time'], '%H:%M')
        datetime.datetime.strptime(data['departure_time'], '%H:%M')

        # Check if arrival time is greater than or equal to departure time
        if data['arrival_timestamp'] >= data['departure_timestamp']:
            return False, 'Arrival time cannot be greater than or equal to departure time.'

        # If all validations pass
        return True, None
    except ValueError:
        # If any parsing fails, return False and an error message
        return False, 'Invalid date or time format.'


def add_timestamp(request):

    train_name = request['train_name']
    train_number = request['train_number']
    arrival_date = request['start_date']
    arrival_time = request['start_time']
    departure_time = request['end_time']
    # next_day = request['next_day'] if  else None
    departure_date = request['next_day_date'] if 'next_day' in request else arrival_date

    # Convert arrival_date and arrival_time to Unix timestamp for arrival
    arrival_datetime = datetime.datetime.strptime(
        arrival_date + ' ' + arrival_time, '%Y-%m-%d %H:%M')
    arrival_timestamp = int(arrival_datetime.timestamp())

    # Convert arrival_date and departure_time to Unix timestamp for departure
    departure_datetime = datetime.datetime.strptime(
        departure_date + ' ' + departure_time, '%Y-%m-%d %H:%M')
    departure_timestamp = int(departure_datetime.timestamp())

    mydict = {
        'train_name': train_name,
        'train_number': train_number,
        'arrival_date': arrival_date,
        'departure_date': departure_date,
        'arrival_time': arrival_time,
        'departure_time': departure_time,
        'arrival_timestamp': arrival_timestamp,
        'departure_timestamp': departure_timestamp
    }
    return mydict


# def remove_id_fields(data):
#     cleaned_data = {}

#     for key, items in data.items():
#         cleaned_items = []

#         for item in items:
#             cleaned_item = {key: value for key, value in item.items() if key not in [
#                 '_id', 'id']}
#             cleaned_items.append(cleaned_item)

#         cleaned_data[key] = cleaned_items

#     return cleaned_data


def export_to_excel(platform_confirm, remaining_activity):
    df = pd.concat([
        pd.DataFrame({
            'Platform': [i] * len(platform_confirm[i]),
            'Train Name': [j['train_name'] for j in platform_confirm[i]],
            'Train Number': [j['train_number'] for j in platform_confirm[i]],
            'Arrival Date': [j['arrival_date'] for j in platform_confirm[i]],
            'Departure Date': [j['departure_date'] for j in platform_confirm[i]],
            'Arrival Time': [j['arrival_time'] for j in platform_confirm[i]],
            'Departure Time': [j['departure_time'] for j in platform_confirm[i]],
        }) for i in platform_confirm.keys()
    ] + [
        pd.DataFrame({
            'Platform': ['Remaining'] * len(remaining_activity),
            'Train Name': [i['train_name'] for i in remaining_activity],
            'Train Number': [i['train_number'] for i in remaining_activity],
            'Arrival Date': [i['arrival_date'] for i in remaining_activity],
            'Departure Date': [i['departure_date'] for i in remaining_activity],
            'Arrival Time': [i['arrival_time'] for i in remaining_activity],
            'Departure Time': [i['departure_time'] for i in remaining_activity],
        })
    ], ignore_index=True)

    # Save the DataFrame to Excel file with styling
    output_path = os.path.join(STATIC_DIR, 'output.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Define a pattern fill for the cells
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00",
                                 end_color="00FF00", fill_type="solid")

        # Apply the pattern fill to the specific cells
        for row_num in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_num, column=1)

            # Check if the value is 'Remaining'
            if cell.value == 'Remaining':
                # Iterate through each cell in the row and apply the yellow fill
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = yellow_fill
            else:
                # If not 'Remaining', fill the entire row with green
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = green_fill

    return True


def get_all_trains():
    """Get all the trains from the database."""

    data = list(collection.find({},id_filter))
    return data